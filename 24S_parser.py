import os
import requests
import json
import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException

# Инициализация Chrome WebDriver
chrome_options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=chrome_options)
last_scraped_page = 0
savepoint_file = 'savepoint.txt'

# Загрузка уже сделанных товаров из Excel
def load_scraped_urls(excel_file):
    if os.path.exists(excel_file):
        df = pd.read_excel(excel_file)
        return df['Product URL'].tolist()
    return []


def save_parsed_urls_to_json(parsed_urls, filename='parsed_urls.json'):
    with open(filename, 'w') as file:
        json.dump(parsed_urls, file)

def load_parsed_urls_from_json(filename='parsed_urls.json'):
    try:
        with open(filename, 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return []

def handle_popup_windows():
    try:
        # Обработка первого всплывающего окна (для нажатия кнопки "Принять все")
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'onetrust-accept-btn-handler'))).click()
        time.sleep(1)

        # Проверка наличия второго окна
        if len(driver.window_handles) > 1:
            # Обработка второго всплывающего окна (для выбора Америки)
            driver.switch_to.window(driver.window_handles[1])
            WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, 'btn-current-country'))).click()
            time.sleep(1)

            # Возврат к главному окну
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(1)

    except Exception as e:
        print(f"Error in handle_popup_windows: {e}")

def get_last_savepoint():
    if os.path.exists(savepoint_file):
        with open(savepoint_file, 'r') as file:
            return int(file.read().strip())
    return 0  # Начать с начала, если не существует точки сохранения

# Функция установки точки сохранения
def set_savepoint(index):
    with open(savepoint_file, 'w') as file:
        file.write(str(index))

def download_image(url, output_directory="downloaded_images"):
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
        
    filename = os.path.join(output_directory, url.split("/")[-1].split(".")[0] + ".jpg")
    if not os.path.exists(filename):
        response = requests.get(url, stream=True)
        with open(filename, 'wb') as file:
            for chunk in response.iter_content(chunk_size=8192):
                file.write(chunk)
    return os.path.basename(filename)



def download_images_for_product():
    # Инициализация набора для исключения дублирования URL-адресов изображений
    image_urls = set()

    # Проверка наличия кнопки "Далее"
    next_buttons = driver.find_elements(By.CSS_SELECTOR, "button.slick-next")

    if not next_buttons:  # Если кнопка "Далее" не существует
        # Получение видимого изображения с разрешением 555x625
        visible_image = driver.find_element(By.CSS_SELECTOR, "div.slick-slide.slick-active.slick-current img[src*='555x625']")
        image_urls.add(visible_image.get_attribute("src"))
    else:
        next_button = next_buttons[0]
        # Если найдено, нажмите кнопку "Далее" до 5 раз
        for _ in range(5):
            images = driver.find_elements(By.CSS_SELECTOR, "div.slick-track img[src]")
            for image in images:
                image_urls.add(image.get_attribute("src"))
            
            # Нажмите кнопку "Далее", чтобы увидеть другие изображения
            try:
                next_button.click()
                time.sleep(2)  # Дать время на загрузку изображений после щелчка мыши
            except ElementClickInterceptedException:
                break

    # Отфильтровать изображения, размер которых равен 555x625
    filtered_urls = {url for url in image_urls if '555x625' in url}
    
    downloaded_filenames = []
    for url in filtered_urls:
        filename = download_image(url)
        if filename:
            downloaded_filenames.append(filename)
    return downloaded_filenames


def click_color_button(button, color_name):
    try:
        # Прокрутка элемента в поле зрения
        driver.execute_script("arguments[0].scrollIntoView(true);", button)
        time.sleep(2)
        
       # Подождать, пока кнопка станет кликабельной
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, f'//img[@alt="{color_name}"]/ancestor::button')))
        
       # Попытка нажать на кнопку
        button.click()
    except NoSuchElementException:
        print(f"Element with color name '{color_name}' not found.")
    except:
       # Если стандартный щелчок не удался, попытается выполнить щелчок на JS
        driver.execute_script("arguments[0].click();", button)




def extract_product_data_from_soup(soup, product_url, image_file_names=[]):
    product_name = soup.find('span', {'data-cy': 'pdp-product-name-text'}).get_text(strip=True) if soup.find('span', {'data-cy': 'pdp-product-name-text'}) else None
    brand = soup.find('a', {'data-cy': 'pdp-brand-anchor'}).get_text(strip=True) if soup.find('a', {'data-cy': 'pdp-brand-anchor'}) else None

    categories = [a.get_text(strip=True) for a in soup.select('.breadcrumb li a')]
    categories = categories[1:]
    while len(categories) < 4:
        categories.append(None)

    # Извлечение информации о продукте
    accordion = soup.find('div', class_='accordion-text')
    
    # Инициализация переменных для хранения извлеченной информации
    description = None
    color = None
    material = None
    size_measurement = None
    country_of_manufacture = None

    if accordion:
        lis = accordion.find_all('li')
        for li in lis:
            title = li.find('span', class_='t-desc accordion-list-item-title')
            if title:
                title_text = title.text.strip().lower()

                if 'description' in title_text:
                    # Извлечение описания
                    desc_spans = li.find_all('span', class_='accordion-list-item-text')
                    description = ' '.join(span.get_text(strip=True) for span in desc_spans)
                elif 'material' in title_text:
                    material = li.text.replace(title.text, '').strip()
                elif 'color' in title_text:
                    # Извлечение информации о цвете и удаление нежелательных символов
                    color_content = li.text.split(':')[-1].strip().replace('_', ' ')
                    if color_content: 
                        color = color_content
                elif 'size & measurements' in title_text:
                    size_measurement = li.text.replace(title.text, '').strip().split('View size guide')[0]
                elif 'country of manufacture' in title_text:
                    country_of_manufacture = li.text.replace(title.text, '').strip()

    # Если цвет не задан, попробробовать извлечь его из описания
    if not color:
        color_description = soup.find('span', string='Color')
        if color_description:
            color = color_description.find_next_sibling('span').get_text(strip=True)
        else:
            color = 'N/A'

    return {
        'Product Name': product_name,
        'Brand': brand,
        'Product URL': product_url,
        'Category 1': categories[0],
        'Category 2': categories[1],
        'Category 3': categories[2],
        'Category 4': categories[3],
        'Description': description,
        'Color': color,
        'Material': material,
        'Size & Measurements': size_measurement,
        'Country of Manufacture': country_of_manufacture,
        'Images': ", ".join(image_file_names)
    }

MAX_RETRIES = 3

def scrape_product_information(product_url, retries=MAX_RETRIES):
    try:
        driver.get(product_url)
        time.sleep(2)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
    
        color_elements = soup.select('ul[data-cy="pdp-color-selector"] li button span picture img')

        all_product_data = []

        if color_elements:
            for color_element in color_elements:
                color_name = color_element['alt']
                try:
                    button = driver.find_element(By.XPATH, f'//img[@alt="{color_name}"]/ancestor::button')
                    button.click()
                    time.sleep(2)

                   # Загрузить изображения для этого конкретного цвета
                    image_file_names = download_images_for_product()

                    # Получить обновленный URL
                    updated_product_url = driver.current_url

                    # Продолжить извлечение данных
                    soup_updated = BeautifulSoup(driver.page_source, 'html.parser')
                    product_data = extract_product_data_from_soup(soup_updated, updated_product_url, image_file_names)

                    if product_data not in all_product_data:
                        all_product_data.append(product_data)
                except NoSuchElementException:
                    print(f"Color element '{color_name}' not found on the page. Skipping this color.")
                    continue
        else: # Работа с продуктами, не имеющими переключателя цветов
            image_file_names = download_images_for_product()
        
            product_data = extract_product_data_from_soup(soup, product_url, image_file_names)
            if not product_data['Color']:
                product_data['Color'] = 'N/A'
            all_product_data.append(product_data)

        return all_product_data

    except Exception as e:
        if retries > 0:
            print(f"Error encountered: {str(e)}. Retrying ({retries} retries left)...")
            time.sleep(5) 
            return scrape_product_information(product_url, retries - 1)
        else:
            print(f"Failed to scrape information for {product_url} after {MAX_RETRIES} attempts.")
            return []

def append_product_to_excel(output_filename, product_data):
    if not os.path.exists(output_filename):
        workbook = Workbook()
        sheet = workbook.active
        headers = [
            'Product Name',
            'Brand',
            'Product URL',
            'Category 1',
            'Category 2',
            'Category 3',
            'Category 4',
            'Description',
            'Color',
            'Material',
            'Size & Measurements',
            'Country of Manufacture',
            'Images'
        ]
        sheet.append(headers)
    else:
        workbook = load_workbook(output_filename)
        sheet = workbook.active

    for data in product_data:
        sheet.append([
            data['Product Name'],
            data['Brand'],
            data['Product URL'],
            data['Category 1'],
            data['Category 2'],
            data['Category 3'],
            data['Category 4'],
            data['Description'],
            data['Color'],
            data['Material'],
            data['Size & Measurements'],
            data['Country of Manufacture'],
            data['Images']
        ])
    workbook.save(output_filename)

def scrape_page(page_url, output_filename, image_output_directory, downloaded_images, already_scraped_urls):
    driver.get(page_url)
    time.sleep(2)
    soup = BeautifulSoup(driver.page_source, 'html.parser')

    handle_popup_windows()

    product_urls = []
    product_elements = soup.find_all('a', {'class': 'product_btn__QSoXG'}, href=True)
    for product_element in product_elements:
        product_url = 'https://www.24s.com' + product_element['href']
        if product_url not in already_scraped_urls:  # Добавлять только новые URL-адреса продуктов
            product_urls.append(product_url)

    for product_url in product_urls:
        all_product_data = scrape_product_information(product_url)
        if all_product_data:  # Добавлять только при наличии данных
            append_product_to_excel(output_filename, all_product_data)



def scrape_website(url, output_filename, image_output_directory):
    driver.get(url)
    time.sleep(2)
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    total_pages = int(soup.find('span', {'class': 'pagination_pageOf__vmVQq'}).text.split()[-1])

    downloaded_images = set()

    for page in range(get_last_savepoint() + 1, total_pages + 1):
        page_url = f"{url}?page={page}"
        print(f"Scraping page {page} of {total_pages}")
        scrape_page(page_url, output_filename, image_output_directory, downloaded_images, already_scraped_urls)
        set_savepoint(page)


if __name__ == '__main__':
    website_url = "https://www.24s.com/en-us/women/ready-to-wear"
    output_filename = "24s_products.xlsx"
    image_output_directory = "images"

    already_scraped_urls = load_scraped_urls(output_filename)

    if not os.path.exists(image_output_directory):
        os.makedirs(image_output_directory)

    if not os.path.exists(output_filename):
        workbook = Workbook()
        sheet = workbook.active
        headers = [
            'Product Name',
            'Brand',
            'Product URL',
            'Category 1',
            'Category 2',
            'Category 3',
            'Category 4',
            'Description',
            'Color',
            'Material',
            'Size & Measurements',
            'Country of Manufacture',
            'Images'
        ]
        sheet.append(headers)
        workbook.save(output_filename)

    # Механизм повторных попыток
    max_retries = 10 
    retry_count = 0
    while retry_count < max_retries:
        try:
            scrape_website(website_url, output_filename, image_output_directory)
            driver.quit()
            break
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            print("Please check the error and decide if you want to continue or stop.")
            input("Press Enter to continue the script or Ctrl+C to stop...")
        print("All retries failed.")
        input("Press Enter to close the script...")

